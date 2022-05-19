import openpyxl
from datetime import datetime
import sys
from cadastro_vendas import *
from PyQt5.QtWidgets import QMainWindow, QApplication
from openpyxl.styles.borders import Border, Side


class Novo(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)
        self.btnSalvar.clicked.connect(self.salvar)
        self.btnNovo.clicked.connect(self.nova_venda)
        self.btnIncluirProduto.clicked.connect(self.incluir_produto)

# salvando no excel

    def salvar(self):

        pedidos = openpyxl.load_workbook(r'C:\Users\Ribas\Documents\Vegas\2022 CONTROLE CAIXA VEGAS.xlsx')
        # nome_planilhas = pedidos.sheetnames
        planilha1 = pedidos[pedidos.sheetnames[0]]
        dt = datetime.now().strftime('%d/%m')

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for linha in range(5, 9999):
            if planilha1.cell(linha, 1).value is None:

                planilha1.cell(linha, 1).value = dt
                planilha1.cell(linha, 2).value = str(self.inputNome.text().title())
                planilha1.cell(linha, 3).value = str(self.inputBairro.text().title())
                planilha1.cell(linha, 4).value = str(self.cbProdutos.currentText())
                planilha1.cell(linha, 5).value = str(self.cbHamburguer.currentText())
                if self.inputQuantidade.text() == '':
                    self.msg.setText(
                        str('<span style="color: #ff0000; font-size: 18px;">Quantidade não pode ser 0.'))
                    break
                else:
                    planilha1.cell(linha, 6).value = int(self.inputQuantidade.text())
                planilha1.cell(linha, 7).value = str(self.cbTipoDaVenda.currentText())
                if self.inputValor.text() == '':
                    self.msg.setText(
                        str('<span style="color: #ff0000; font-size: 18px;">Valor não pode ser 0.'))
                    break
                else:
                    planilha1.cell(linha, 8).value = float(self.inputValor.text().replace(',', '.'))
                planilha1.cell(linha, 9).value = str(self.cbFormaDePagamento.currentText())
                if self.inputTaxaDeEntrega.text() == '':
                    planilha1.cell(linha, 10).value = float(0)
                else:
                    planilha1.cell(linha, 10).value = float(self.inputTaxaDeEntrega.text().replace(',', '.'))

                planilha1.cell(linha, 1).border = thin_border
                planilha1.cell(linha, 2).border = thin_border
                planilha1.cell(linha, 3).border = thin_border
                planilha1.cell(linha, 4).border = thin_border
                planilha1.cell(linha, 5).border = thin_border
                planilha1.cell(linha, 6).border = thin_border
                planilha1.cell(linha, 7).border = thin_border
                planilha1.cell(linha, 8).border = thin_border
                planilha1.cell(linha, 9).border = thin_border
                planilha1.cell(linha, 10).border = thin_border

                self.msg.setText(str('<span style="color: #008000; font-size: 18px;">Venda registrada com sucesso.'))
                break
        pedidos.save(r'C:\Users\Ribas\Documents\Vegas\2022 CONTROLE CAIXA VEGAS.xlsx')

    def nova_venda(self):

        self.inputNome.setText(str(''))
        self.inputBairro.setText(str(''))
        self.cbProdutos.setCurrentText(str(''))
        self.cbHamburguer.setCurrentText(str(''))
        self.inputQuantidade.setText(str(''))
        self.cbTipoDaVenda.setCurrentText(str(''))
        self.inputValor.setText(str(''))
        self.cbFormaDePagamento.setCurrentText(str(''))
        self.inputTaxaDeEntrega.setText(str(''))
        self.msg.setText(str(''))

    def incluir_produto(self):

        self.cbProdutos.setCurrentText(str(''))
        self.cbHamburguer.setCurrentText(str(''))
        self.inputQuantidade.setText(str(''))
        self.inputValor.setText(str(''))
        self.inputTaxaDeEntrega.setText(str(''))
        self.msg.setText(str(''))


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    novo = Novo()
    novo.show()
    qt.exec_()
